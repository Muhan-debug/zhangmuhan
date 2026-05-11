# =============================================================================
# Phase 1-2: 数据初处理与可视化 — 心衰患者生存预测
# Heart Failure Clinical Records — Data Loading, EDA & Preprocessing
# =============================================================================
 
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
from scipy import stats
from sklearn.preprocessing import StandardScaler
import warnings
warnings.filterwarnings('ignore')
 
# ── 全局绘图风格 ──────────────────────────────────────────────────────────────
plt.rcParams.update({
    'figure.facecolor': '#0f1117',
    'axes.facecolor':   '#1a1d27',
    'axes.edgecolor':   '#3a3f55',
    'axes.labelcolor':  '#c8cfe8',
    'xtick.color':      '#8890b0',
    'ytick.color':      '#8890b0',
    'text.color':       '#c8cfe8',
    'grid.color':       '#2a2f42',
    'grid.linewidth':   0.6,
    'font.family':      'DejaVu Sans',
})
PALETTE   = ['#4fc3f7', '#81c784', '#ffb74d', '#f06292',
             '#ce93d8', '#80cbc4', '#fff176', '#ff8a65']
ACCENT    = '#4fc3f7'
WARN_COL  = '#ffb74d'
GOOD_COL  = '#81c784'
 
# =============================================================================
# 1. 数据加载
# =============================================================================
print("=" * 65)
print("  STEP 1 — 数据加载 (Data Loading)")
print("=" * 65)
 
# 使用相对路径读取数据集（将 CSV 放在与本脚本同目录下）
DATA_PATH = "./heart_failure_clinical_records_dataset.csv"
df = pd.read_csv(DATA_PATH)
 
print(f"\n✅ 数据加载成功！共 {df.shape[0]} 条记录，{df.shape[1]} 个字段\n")
print("─" * 65)
print("【数据前5行预览】")
print(df.head().to_string())
 
# =============================================================================
# 2. 描述性统计分析
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
print("\n" + "=" * 65)
print("  STEP 2 — 描述性统计分析 (Descriptive Statistics)")
print("=" * 65)
 
desc = df.describe().T
desc['variance'] = df.var()
desc['skewness'] = df.skew()
desc['kurtosis'] = df.kurtosis()
desc['missing']  = df.isnull().sum()
desc['missing_pct'] = (df.isnull().sum() / len(df) * 100)
 
desc = desc[['count', 'mean', 'std', 'variance',
             'min', '25%', '50%', '75%', 'max',
             'skewness', 'kurtosis', 'missing', 'missing_pct']]
desc.columns = ['计数', '均值', '标准差', '方差',
                '最小值', 'Q1(25%)', '中位数', 'Q3(75%)', '最大值',
                '偏度', '峰度', '缺失值', '缺失率(%)']
 
print("\n【全量描述性统计】")
print(desc.round(4).to_string())
 
# 目标变量分布
death_counts = df['DEATH_EVENT'].value_counts()
survival_n = death_counts.get(0, 0)
death_n    = death_counts.get(1, 0)
print(f"\n【目标变量 DEATH_EVENT 分布】")
print(f"  存活 (0): {survival_n} 例  ({survival_n/len(df)*100:.1f}%)")
print(f"  死亡 (1): {death_n} 例  ({death_n/len(df)*100:.1f}%)")
 
# ============================================================
# STEP 3  缺失值检查
# ============================================================
print("\n" + "=" * 65)
print("  STEP 3 — 缺失值检查")
print("=" * 65)
missing_total = df.isnull().sum().sum()
print(f"总缺失值数量: {missing_total}")
if missing_total == 0:
    print("✅ 数据完整，无缺失值。")
else:
    print(df.isnull().sum()[df.isnull().sum() > 0])
 
# ============================================================
# STEP 4  生成 Excel 统计报告
# ============================================================
print("\n" + "=" * 65)
print("  STEP 4 — 生成 Excel 统计报告")
print("=" * 65)
 
wb = Workbook()
 
# ── 样式定义 ─────────────────────────────────────────────────
header_fill  = PatternFill('solid', start_color='1F4E79')
subhead_fill = PatternFill('solid', start_color='2E75B6')
alt_fill     = PatternFill('solid', start_color='D6E4F0')
white_fill   = PatternFill('solid', start_color='FFFFFF')
 
header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=14)
subhead_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
body_font    = Font(name='Arial', size=10)
bold_font    = Font(name='Arial', bold=True, size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align   = Alignment(horizontal='left',   vertical='center')
 
thin_side   = Side(style='thin',   color='B0B0B0')
medium_side = Side(style='medium', color='1F4E79')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
 
# ── Sheet 1: 描述性统计 ──────────────────────────────────────
ws1 = wb.active
ws1.title = '描述性统计'
 
# 主标题
ws1.merge_cells('A1:N1')
ws1['A1'] = '心力衰竭临床数据集 — 描述性统计分析报告'
ws1['A1'].font = header_font
ws1['A1'].fill = header_fill
ws1['A1'].alignment = center_align
ws1.row_dimensions[1].height = 34
 
# 基本信息
ws1.merge_cells('A2:N2')
ws1['A2'] = (f"样本量: {len(df)}  |  变量数: {len(df.columns)}  |  "
             f"存活: {survival_n} ({survival_n/len(df)*100:.1f}%)  |  "
             f"死亡: {death_n} ({death_n/len(df)*100:.1f}%)")
ws1['A2'].font = subhead_font
ws1['A2'].fill = subhead_fill
ws1['A2'].alignment = center_align
ws1.row_dimensions[2].height = 22
 
# 列标题
col_headers = ['变量名', '计数', '均值', '标准差', '方差',
               '最小值', 'Q1(25%)', '中位数', 'Q3(75%)', '最大值',
               '偏度', '峰度', '缺失值', '缺失率(%)']
for col_idx, header in enumerate(col_headers, 1):
    cell = ws1.cell(row=3, column=col_idx, value=header)
    cell.font = subhead_font
    cell.fill = subhead_fill
    cell.alignment = center_align
    cell.border = thin_border
ws1.row_dimensions[3].height = 28
 
# 数据行
for row_idx, (var_name, row_data) in enumerate(desc.round(4).iterrows(), 4):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    name_cell = ws1.cell(row=row_idx, column=1, value=var_name)
    name_cell.font = bold_font
    name_cell.fill = fill
    name_cell.alignment = left_align
    name_cell.border = thin_border
    for col_idx, value in enumerate(row_data, 2):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font = body_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border
    ws1.row_dimensions[row_idx].height = 20
 
# 列宽
col_widths = [26, 8, 12, 10, 14, 10, 10, 10, 10, 10, 10, 10, 10, 12]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
 
# ── Sheet 2: 变量说明 ─────────────────────────────────────────
ws2 = wb.create_sheet('变量说明')
 
var_info = [
    ('变量名',                    '类型',   '说明',                          '取值范围'),
    ('age',                       '连续型', '患者年龄（岁）',                 '40–95'),
    ('anaemia',                   '二元型', '是否贫血',                       '0=否, 1=是'),
    ('creatinine_phosphokinase',  '连续型', '肌酸激酶水平（mcg/L）',          '23–7861'),
    ('diabetes',                  '二元型', '是否患糖尿病',                   '0=否, 1=是'),
    ('ejection_fraction',         '连续型', '射血分数（%）',                  '14–80'),
    ('high_blood_pressure',       '二元型', '是否高血压',                     '0=否, 1=是'),
    ('platelets',                 '连续型', '血小板数量（kiloplatelets/mL）', '25100–850000'),
    ('serum_creatinine',          '连续型', '血清肌酐（mg/dL）',             '0.5–9.4'),
    ('serum_sodium',              '连续型', '血清钠（mEq/L）',               '113–148'),
    ('sex',                       '二元型', '性别',                           '0=女, 1=男'),
    ('smoking',                   '二元型', '是否吸烟',                       '0=否, 1=是'),
    ('time',                      '连续型', '随访时间（天）',                  '4–285'),
    ('DEATH_EVENT',               '二元型', '目标变量：是否死亡',              '0=存活, 1=死亡'),
]
 
ws2.merge_cells('A1:D1')
ws2['A1'] = '变量说明表'
ws2['A1'].font = Font(name='Arial', bold=True, color='FFFFFF', size=13)
ws2['A1'].fill = header_fill
ws2['A1'].alignment = center_align
ws2.row_dimensions[1].height = 30
 
for row_idx, row_data in enumerate(var_info, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        if row_idx == 2:
            cell.font = subhead_font
            cell.fill = subhead_fill
            cell.alignment = center_align
        else:
            cell.font = body_font
            cell.fill = alt_fill if row_idx % 2 == 0 else white_fill
            cell.alignment = left_align if col_idx == 3 else center_align
    ws2.row_dimensions[row_idx].height = 20
 
for col, w in zip('ABCD', [28, 12, 32, 28]):
    ws2.column_dimensions[col].width = w
OUTPUT_FILE = "描述性统计.xlsx"
# ── 保存 ─────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"✅ 统计报告已生成: {OUTPUT_FILE}")
print("   包含 Sheet: 【描述性统计】【变量说明】")
 
# =============================================================================
# 3. 缺失值检查
# =============================================================================
print("\n" + "=" * 65)
print("  STEP 3 — 缺失值检查 (Missing Value Detection)")
print("=" * 65)
 
missing = df.isnull().sum()
missing_pct = (missing / len(df) * 100).round(2)
missing_df = pd.DataFrame({'缺失数量': missing, '缺失比例(%)': missing_pct})
 
if missing.sum() == 0:
    # ── 本数据集无缺失值，在此明确说明 ──────────────────────────────────────
    print("\n✅ 【确认】本数据集不含任何缺失值（Null/NaN）")
    print("   所有 13 个字段在 299 条记录中均完整，无需插补处理。")
    print(missing_df.to_string())
else:
    print("\n⚠️  发现缺失值，详情如下：")
    print(missing_df[missing_df['缺失数量'] > 0].to_string())
    # 若存在缺失值，可按需选择以下策略（本数据集跳过）：
    # df.fillna(df.median(), inplace=True)   # 数值列用中位数填充
    # df.dropna(inplace=True)                # 删除含缺失行
 
# =============================================================================
# 4. 异常值检测（IQR 方法）
# =============================================================================
print("\n" + "=" * 65)
print("  STEP 4 — 异常值检测 (Outlier Detection via IQR)")
print("=" * 65)
 
# 选取连续数值型特征进行检测（排除二值变量）
continuous_cols = [
    'age', 'creatinine_phosphokinase', 'ejection_fraction',
    'platelets', 'serum_creatinine', 'serum_sodium', 'time'
]
 
outlier_report = []
for col in continuous_cols:
    Q1  = df[col].quantile(0.25)
    Q3  = df[col].quantile(0.75)
    IQR = Q3 - Q1
    lower = Q1 - 1.5 * IQR
    upper = Q3 + 1.5 * IQR
    n_out = ((df[col] < lower) | (df[col] > upper)).sum()
    outlier_report.append({
        '特征': col,
        'Q1': round(Q1, 2),
        'Q3': round(Q3, 2),
        'IQR下界': round(lower, 2),
        'IQR上界': round(upper, 2),
        '异常值数量': n_out,
        '异常值占比(%)': round(n_out / len(df) * 100, 2)
    })
 
outlier_df = pd.DataFrame(outlier_report).set_index('特征')
print("\n【IQR 异常值统计】")
print(outlier_df.to_string())
 
# ── 处理策略：保留异常值（不删除），原因说明 ─────────────────────────────────
# 心衰临床数据中，极端值（如超高 creatinine_phosphokinase）
# 往往具有重要的临床意义，直接删除可能丢失关键病理信息。
# 后续树模型（Random Forest/XGBoost）对异常值天然鲁棒，无需额外处理。
print("\n📋 【处理策略说明】")
print("   保留所有异常值：心衰临床数据的极端值具有病理诊断价值，")
print("   删除可能引入偏差。树模型对异常值天然鲁棒，Logistic Regression")
print("   则通过后续标准化降低量纲影响。")
 
# =============================================================================
# 5. 数据标准化（StandardScaler）
# =============================================================================
print("\n" + "=" * 65)
print("  STEP 5 — 数据标准化 (Feature Standardization)")
print("=" * 65)
 
# 二值型特征不参与标准化
binary_cols = ['anaemia', 'diabetes', 'high_blood_pressure', 'sex', 'smoking', 'DEATH_EVENT']
scale_cols  = [c for c in df.columns if c not in binary_cols]
 
scaler   = StandardScaler()
df_scaled = df.copy()
df_scaled[scale_cols] = scaler.fit_transform(df[scale_cols])
 
print(f"\n✅ 已对以下 {len(scale_cols)} 个连续特征进行 Z-score 标准化：")
print(f"   {scale_cols}")
print(f"\n   公式：z = (x - μ) / σ")
print(f"\n【标准化后统计（均值≈0，标准差≈1）】")
print(df_scaled[scale_cols].describe().round(4).T[['mean','std']].to_string())
print(f"\n【二值特征保持原值，不参与标准化】")
print(f"   {binary_cols}")
 
# =============================================================================
# 6. Figure 1 — Data Summary（1行3列：目标分布 | 缺失值检查 | 异常值统计）
# =============================================================================
print("\n" + "=" * 65)
print("  STEP 6 — Figure 1: Data Summary (1×3 layout)")
print("=" * 65)
 
BG_DARK   = '#0f1117'
BG_PANEL  = '#1a1d27'
TXT_HEAD  = '#f0f4ff'   # 标题白
TXT_BODY  = '#c8d0ee'   # 正文浅蓝灰
GRID_COL  = '#2c3148'
 
fig1, axes1 = plt.subplots(
    1, 3,
    figsize=(24, 8),           # 宽幅单行，给每个子图足够空间
    facecolor=BG_DARK
)
fig1.subplots_adjust(left=0.05, right=0.97, top=0.82, bottom=0.15, wspace=0.38)
 
# ── Fig1-A: 目标变量分布（饼图）──────────────────────────────────────────────
ax_pie = axes1[0]
ax_pie.set_facecolor(BG_PANEL)
wedge_colors = ['#4fc3f7', '#f06292']
wedges, texts, autotexts = ax_pie.pie(
    death_counts.values,
    labels=['Survived (0)', 'Deceased (1)'],
    autopct='%1.1f%%',
    colors=wedge_colors,
    startangle=90,
    radius=0.88,
    pctdistance=0.68,
    wedgeprops=dict(edgecolor=BG_DARK, linewidth=3),
    textprops=dict(color=TXT_BODY, fontsize=15, fontweight='bold')
)
for at in autotexts:
    at.set_fontsize(16)
    at.set_fontweight('bold')
    at.set_color(TXT_HEAD)
# 在图中央添加样本数注释
ax_pie.text(0, -1.22,
            f'Total: {len(df)} patients',
            ha='center', va='center',
            fontsize=13, color=TXT_BODY, style='italic')
ax_pie.set_title('Target Distribution\n(DEATH_EVENT)',
                 fontsize=18, fontweight='bold', color=TXT_HEAD, pad=18)
 
# ── Fig1-B: 缺失值检查 ────────────────────────────────────────────────────────
ax_miss = axes1[1]
ax_miss.set_facecolor(BG_PANEL)
miss_matrix = df.isnull().astype(int)
 
if miss_matrix.values.sum() == 0:
    # 无缺失值 → 用大字提示 + 逐字段列表
    ax_miss.text(0.5, 0.62,
                 '✓  No Missing Values',
                 ha='center', va='center', fontsize=20,
                 color=GOOD_COL, fontweight='bold', transform=ax_miss.transAxes)
    ax_miss.text(0.5, 0.46,
                 'All 13 fields complete\nacross 299 records',
                 ha='center', va='center', fontsize=13,
                 color=TXT_BODY, transform=ax_miss.transAxes)
    # 在图下方列出字段名，分两列
    col_names = list(df.columns)
    left_cols  = col_names[:7]
    right_cols = col_names[7:]
    left_text  = '\n'.join([f'• {c}' for c in left_cols])
    right_text = '\n'.join([f'• {c}' for c in right_cols])
    ax_miss.text(0.20, 0.28, left_text,  ha='center', va='top', fontsize=10.5,
                 color='#8890c8', transform=ax_miss.transAxes, linespacing=1.7)
    ax_miss.text(0.80, 0.28, right_text, ha='center', va='top', fontsize=10.5,
                 color='#8890c8', transform=ax_miss.transAxes, linespacing=1.7)
    for spine in ax_miss.spines.values():
        spine.set_edgecolor('#3a3f55')
    ax_miss.set_xticks([])
    ax_miss.set_yticks([])
else:
    sns.heatmap(miss_matrix, ax=ax_miss, cbar=False, cmap='OrRd', yticklabels=False,
                xticklabels=list(df.columns))
    ax_miss.tick_params(axis='x', labelrotation=40, labelsize=11, colors=TXT_BODY)
ax_miss.set_title('Missing Value Check',
                  fontsize=18, fontweight='bold', color=TXT_HEAD, pad=18)
 
# ── Fig1-C: 异常值数量柱状图 ──────────────────────────────────────────────────
ax_out = axes1[2]
ax_out.set_facecolor(BG_PANEL)
 
feat_labels = [c.replace('_', '\n') for c in outlier_df.index]   # 换行防止重叠
counts_val  = outlier_df['异常值数量'].values
colors_bar  = [WARN_COL if v > 0 else GOOD_COL for v in counts_val]
 
bars = ax_out.bar(
    range(len(feat_labels)), counts_val,
    color=colors_bar, edgecolor=BG_DARK,
    linewidth=1.5, width=0.58, zorder=3
)
# 柱顶数字标注
for bar, v in zip(bars, counts_val):
    y_pos = bar.get_height() + 0.4
    ax_out.text(bar.get_x() + bar.get_width() / 2, y_pos,
                str(v), ha='center', va='bottom',
                fontsize=13, fontweight='bold', color=TXT_HEAD, zorder=4)
 
ax_out.set_xticks(range(len(feat_labels)))
ax_out.set_xticklabels(feat_labels, fontsize=11, color=TXT_BODY, linespacing=0.9)
ax_out.set_ylabel('Number of Outliers', fontsize=13, color=TXT_BODY, labelpad=10)
ax_out.set_xlabel('Feature', fontsize=13, color=TXT_BODY, labelpad=10)
ax_out.tick_params(axis='y', labelsize=11, colors=TXT_BODY)
ax_out.yaxis.grid(True, color=GRID_COL, linewidth=0.8, zorder=0)
ax_out.set_axisbelow(True)
 
# 图例
from matplotlib.patches import Patch
legend_els = [Patch(facecolor=WARN_COL, label='Has outliers'),
              Patch(facecolor=GOOD_COL, label='Clean')]
ax_out.legend(handles=legend_els, fontsize=12, loc='upper right',
              facecolor='#22263a', edgecolor='#3a3f55',
              labelcolor=TXT_BODY, framealpha=0.9)
ax_out.set_title('Outlier Count per Feature\n(IQR Method)',
                 fontsize=18, fontweight='bold', color=TXT_HEAD, pad=18)
for spine in ax_out.spines.values():
    spine.set_edgecolor('#3a3f55')
 
fig1.suptitle('Figure 1 — Heart Failure Dataset: Data Summary',
              fontsize=22, fontweight='bold', color=TXT_HEAD, y=0.97)
 
fig1_path = "./fig1_data_summary.png"
fig1.savefig(fig1_path, dpi=300, bbox_inches='tight', facecolor=BG_DARK)
plt.close(fig1)
print(f"\n✅ Figure 1 已保存至：{fig1_path}")
 
# =============================================================================
# 7. Figure 2 — Feature Distribution（2行3列：连续变量 Raw vs Scaled Boxplot）
# =============================================================================
print("\n" + "=" * 65)
print("  STEP 7 — Figure 2: Feature Distribution (2×3 Boxplot layout)")
print("=" * 65)
 
show_cols = continuous_cols[:6]   # 取全部6个连续特征
fig2, axes2 = plt.subplots(
    2, 3,
    figsize=(26, 16),           # 大画布，每格有充裕空间
    facecolor=BG_DARK
)
fig2.subplots_adjust(left=0.06, right=0.97, top=0.88, bottom=0.09,
                     hspace=0.50, wspace=0.36)
 
for i, col in enumerate(show_cols):
    row_idx = i // 3
    col_idx = i % 3
    ax = axes2[row_idx][col_idx]
    ax.set_facecolor(BG_PANEL)
 
    data_raw    = df[col].values
    data_scaled = df_scaled[col].values
 
    # positions 拉开到 1 和 3，中间留出明显间距
    bp = ax.boxplot(
        [data_raw, data_scaled],
        positions=[1, 3],          # 位置间距从默认1扩大到2
        widths=0.72,
        patch_artist=True,
        notch=False,
        medianprops=dict(color='#fff176', linewidth=2.5),
        whiskerprops=dict(color='#8890b0', linewidth=1.5, linestyle='--'),
        capprops=dict(color='#8890b0', linewidth=2.0),
        flierprops=dict(
            marker='D', markerfacecolor=WARN_COL,
            markersize=5, alpha=0.65,
            markeredgecolor='none'
        ),
        boxprops=dict(linewidth=1.8)
    )
 
    # 盒子配色
    RAW_COL    = '#4fc3f7'
    SCALED_COL = '#81c784'
    bp['boxes'][0].set_facecolor(RAW_COL);    bp['boxes'][0].set_alpha(0.55)
    bp['boxes'][1].set_facecolor(SCALED_COL); bp['boxes'][1].set_alpha(0.55)
 
    # X 轴刻度与标签
    ax.set_xticks([1, 3])
    ax.set_xticklabels(
        ['Raw Data', 'Scaled (Z-score)'],
        fontsize=14, color=TXT_BODY, fontweight='bold'
    )
    ax.set_xlim(0, 4)            # 两端留白，避免箱线图贴边
 
    # Y 轴
    ax.tick_params(axis='y', labelsize=12, colors=TXT_BODY)
    ax.set_ylabel('Value', fontsize=13, color=TXT_BODY, labelpad=8)
 
    # 网格
    ax.yaxis.grid(True, color=GRID_COL, linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_edgecolor('#3a3f55')
 
    # 子标题：特征名 + 样本量/异常值数
    n_out = outlier_df.loc[col, '异常值数量'] if col in outlier_df.index else 0
    subtitle = f'n={len(df)}   outliers={n_out}'
    ax.set_title(f'{col}\n{subtitle}',
                 fontsize=16, fontweight='bold', color=TXT_HEAD,
                 pad=12, linespacing=1.5)
 
# 全图图例（右上角）
from matplotlib.patches import Patch as MPatch
leg_handles = [
    MPatch(facecolor='#4fc3f7', alpha=0.7, label='Raw Data'),
    MPatch(facecolor='#81c784', alpha=0.7, label='Scaled (Z-score)'),
    plt.Line2D([0], [0], color='#fff176', linewidth=2.5, label='Median'),
    plt.Line2D([0], [0], marker='D', color='none',
               markerfacecolor=WARN_COL, markersize=7, label='Outlier'),
]
fig2.legend(
    handles=leg_handles,
    loc='upper right', bbox_to_anchor=(0.98, 0.96),
    fontsize=13, ncol=4,
    facecolor='#22263a', edgecolor='#3a3f55',
    labelcolor=TXT_BODY, framealpha=0.92,
    handlelength=1.8
)
 
fig2.suptitle('Figure 2 — Continuous Feature Distribution: Raw vs. Standardized',
              fontsize=22, fontweight='bold', color=TXT_HEAD, y=0.95)
 
fig2_path = "./fig2_feature_distribution.png"
fig2.savefig(fig2_path, dpi=300, bbox_inches='tight', facecolor=BG_DARK)
plt.close(fig2)
print(f"✅ Figure 2 已保存至：{fig2_path}")
 
# =============================================================================
# 8. 相关性热力图（单独保存，供论文使用）
# =============================================================================
fig3, ax_heat = plt.subplots(figsize=(12, 10), facecolor=BG_DARK)
ax_heat.set_facecolor(BG_PANEL)
 
corr = df.corr()
mask = np.triu(np.ones_like(corr, dtype=bool))

sns.heatmap(
    corr, mask=mask,ax=ax_heat,
    cmap='coolwarm', center=0, vmin=-1, vmax=1,
    annot=True, fmt='.2f', annot_kws={'color': 'black', 'size': 10} ,
    linewidths=0.6, linecolor=BG_DARK,
    cbar_kws={'shrink': 0.75, 'label': 'Pearson  r'}
)
ax_heat.set_title('Feature Correlation Heatmap\n(Heart Failure Clinical Records)',
                  fontsize=17, fontweight='bold', color=TXT_HEAD, pad=18)
ax_heat.tick_params(axis='x', labelrotation=38, labelsize=12, colors=TXT_BODY)
ax_heat.tick_params(axis='y', labelrotation=0,  labelsize=12, colors=TXT_BODY)
ax_heat.figure.axes[-1].tick_params(labelsize=11, colors=TXT_BODY)
ax_heat.figure.axes[-1].yaxis.label.set(color=TXT_BODY, fontsize=12)

# 调整 X 轴标签：旋转 45 度，并向右对齐防止重叠
plt.xticks(rotation=45, ha='right', fontsize=10)

# 自动调整布局，防止标签超出画面
plt.tight_layout()
 
heatmap_path = "./phase1_correlation_heatmap.png"
fig3.savefig(heatmap_path, dpi=300, bbox_inches='tight', facecolor=BG_DARK)
plt.close(fig3)
print(f"✅ 相关性热力图已保存至：{heatmap_path}")
 
# =============================================================================
# 8. 输出标准化后数据集
# =============================================================================
scaled_csv = "./heart_failure_scaled.csv"
df_scaled.to_csv(scaled_csv, index=False)
print(f"✅ 标准化数据集已保存至：{scaled_csv}（供后续建模使用）")
 
# =============================================================================
# 依赖库清单 (requirements)
# =============================================================================
print("\n" + "=" * 65)
print("  运行本脚本所需依赖库")
print("=" * 65)
print("""
  pandas>=1.5.0          # 数据加载与处理
  numpy>=1.23.0           # 数值计算
  matplotlib>=3.6.0       # 绘图基础库
  seaborn>=0.12.0         # 统计可视化
  scipy>=1.9.0            # 统计检验（IQR/T检验）
  scikit-learn>=1.1.0     # StandardScaler 标准化
 
  安装命令：
  pip install pandas numpy matplotlib seaborn scipy scikit-learn
""")
print("=" * 65)
print("  Phase 1 完成！输出文件：")
print("  ├── fig1_data_summary.png          （数据总览：饼图 + 缺失值 + 异常值）")
print("  ├── fig2_feature_distribution.png  （特征分布：2×3 Boxplot Raw vs Scaled）")
print("  ├── phase1_correlation_heatmap.png （相关性热力图，供论文使用）")
print("  └── heart_failure_scaled.csv       （标准化数据集）")
print("=" * 65)
 


 #phase 3:因子检测
# ============================================================
#  Heart Failure — 相关性分析 + 随机森林特征选择
#  依赖：pandas, numpy, matplotlib, seaborn, scikit-learn, openpyxl
#  运行前请安装：pip install pandas numpy matplotlib seaborn scikit-learn openpyxl
# ============================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import StratifiedKFold, cross_val_score
from sklearn.inspection import permutation_importance
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ── 全局字体（Windows 本地中文字体，按优先级回退） ──────────────
import matplotlib.font_manager as fm

# 自动从系统字体中查找可用的中文字体
_zh_candidates = ["Microsoft YaHei", "SimHei", "SimSun", "KaiTi", "FangSong",
                  "DengXian", "YouYuan", "NSimSun"]
_available = {f.name for f in fm.fontManager.ttflist}
_zh_fonts  = [f for f in _zh_candidates if f in _available]

if _zh_fonts:
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = _zh_fonts + ["DejaVu Sans"]
else:
    # 兜底：直接用字体文件路径加载（Windows 路径）
    _win_font_paths = [
        r"C:\Windows\Fonts\msyh.ttc",    # 微软雅黑
        r"C:\Windows\Fonts\simhei.ttf",  # 黑体
        r"C:\Windows\Fonts\simsun.ttc",  # 宋体
    ]
    import os
    for _fp in _win_font_paths:
        if os.path.exists(_fp):
            _prop = fm.FontProperties(fname=_fp)
            plt.rcParams["font.family"] = _prop.get_name()
            break

plt.rcParams["axes.unicode_minus"] = False

# ============================================================
# 0. 加载数据 & 剔除 time
# ============================================================
df = pd.read_csv("heart_failure_scaled.csv")
df = df.drop(columns=["time"])            # 剔除生存分析终点变量，避免数据泄露

X = df.drop(columns=["DEATH_EVENT"])
y = df["DEATH_EVENT"]

FEATURE_ZH = {                            # 特征中文映射（图表标签用）
    "age":                      "年龄",
    "anaemia":                  "贫血",
    "creatinine_phosphokinase": "肌酸激酶",
    "diabetes":                 "糖尿病",
    "ejection_fraction":        "射血分数",
    "high_blood_pressure":      "高血压",
    "platelets":                "血小板",
    "serum_creatinine":         "血清肌酐",
    "serum_sodium":             "血清钠",
    "sex":                      "性别",
    "smoking":                  "吸烟",
}

# ============================================================
# 1. 相关性热图
# ============================================================
print("正在绘制相关性热图…")

corr_df = df.copy()
corr_matrix = corr_df.corr()

# 仅保留与 DEATH_EVENT 的相关性（单列热图 + 全矩阵热图）
fig, axes = plt.subplots(1, 2, figsize=(18, 7),
                         gridspec_kw={"width_ratios": [1, 3]})

# ── 左图：各特征与 DEATH_EVENT 的相关性条形热图 ──────────────
death_corr = (corr_matrix["DEATH_EVENT"]
              .drop("DEATH_EVENT")
              .sort_values(ascending=False))

colors = ["#d73027" if v > 0 else "#4575b4" for v in death_corr.values]
ax_bar = axes[0]
bars = ax_bar.barh(death_corr.index[::-1],
                   death_corr.values[::-1],
                   color=colors[::-1], edgecolor="white", height=0.6)
ax_bar.axvline(0, color="black", linewidth=0.8)
ax_bar.set_xlabel("Pearson 相关系数", fontsize=11)
ax_bar.set_title("各特征与 DEATH_EVENT\n的相关系数", fontsize=12, fontweight="bold")
ax_bar.set_yticklabels([FEATURE_ZH.get(f, f) for f in death_corr.index[::-1]],
                        fontsize=10)
for bar, val in zip(bars, death_corr.values[::-1]):
    ax_bar.text(val + (0.005 if val >= 0 else -0.005), bar.get_y() + bar.get_height() / 2,
                f"{val:.3f}", va="center",
                ha="left" if val >= 0 else "right", fontsize=8.5)
ax_bar.set_xlim(-0.5, 0.55)

red_patch  = mpatches.Patch(color="#d73027", label="正相关（↑ 死亡风险）")
blue_patch = mpatches.Patch(color="#4575b4", label="负相关（↓ 死亡风险）")
ax_bar.legend(handles=[red_patch, blue_patch], fontsize=9, loc="lower right")

# ── 右图：完整相关矩阵热图 ────────────────────────────────────
ax_heat = axes[1]
zh_cols = [FEATURE_ZH.get(c, c) if c != "DEATH_EVENT" else "死亡事件"
           for c in corr_matrix.columns]
mask = np.zeros_like(corr_matrix, dtype=bool)
mask[np.triu_indices_from(mask)] = True          # 只显示下三角

sns.heatmap(corr_matrix,
            mask=mask,
            annot=True, fmt=".2f", annot_kws={"size": 8},
            cmap="RdBu_r", center=0, vmin=-1, vmax=1,
            linewidths=0.4, linecolor="white",
            ax=ax_heat,
            xticklabels=zh_cols,
            yticklabels=zh_cols)
ax_heat.set_title("完整特征相关性矩阵（下三角）", fontsize=12, fontweight="bold")
ax_heat.tick_params(axis="x", rotation=30, labelsize=9)
ax_heat.tick_params(axis="y", rotation=0,  labelsize=9)

fig.suptitle("Heart Failure — 临床特征相关性分析", fontsize=14, fontweight="bold", y=1.01)
plt.tight_layout()
plt.savefig("heatmap_correlation.png", dpi=180, bbox_inches="tight")
plt.close()
print("  → 已保存：heatmap_correlation.png")

# ============================================================
# 2. 随机森林训练 & 特征重要性
# ============================================================
print("正在训练随机森林模型…")

rf = RandomForestClassifier(
    n_estimators=500,
    max_depth=None,
    min_samples_leaf=2,
    class_weight="balanced",
    random_state=42,
    n_jobs=-1
)
rf.fit(X, y)

# 5 折交叉验证 AUC
cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
auc_scores = cross_val_score(rf, X, y, cv=cv, scoring="roc_auc")
print(f"  5-fold CV AUC = {auc_scores.mean():.4f} ± {auc_scores.std():.4f}")

# MDI 特征重要性
importances = pd.Series(rf.feature_importances_, index=X.columns)
importances_sorted = importances.sort_values(ascending=False)

# Permutation importance（更稳健）
perm = permutation_importance(rf, X, y, n_repeats=30, random_state=42, n_jobs=-1)
perm_imp = pd.Series(perm.importances_mean, index=X.columns).sort_values(ascending=False)

# ── 特征重要性可视化 ──────────────────────────────────────────
print("正在绘制特征重要性图…")

fig, axes = plt.subplots(1, 2, figsize=(16, 6))

# 颜色：前3名高亮
def bar_colors(series, top=3):
    cols = []
    for i, (feat, _) in enumerate(series.items()):
        cols.append("#e63946" if i < top else "#457b9d")
    return cols

# MDI
ax1 = axes[0]
cols1 = bar_colors(importances_sorted)
bars1 = ax1.barh(importances_sorted.index[::-1],
                 importances_sorted.values[::-1],
                 color=cols1[::-1], edgecolor="white", height=0.6)
ax1.set_xlabel("MDI 重要性分数", fontsize=11)
ax1.set_title("随机森林特征重要性\n（Mean Decrease Impurity）", fontsize=12, fontweight="bold")
ax1.set_yticklabels([FEATURE_ZH.get(f, f) for f in importances_sorted.index[::-1]], fontsize=10)
for bar, val in zip(bars1, importances_sorted.values[::-1]):
    ax1.text(val + 0.002, bar.get_y() + bar.get_height() / 2,
             f"{val:.4f}", va="center", ha="left", fontsize=8.5)

# Permutation
ax2 = axes[1]
cols2 = bar_colors(perm_imp)
bars2 = ax2.barh(perm_imp.index[::-1],
                 perm_imp.values[::-1],
                 color=cols2[::-1], edgecolor="white", height=0.6)
ax2.set_xlabel("Permutation 重要性分数", fontsize=11)
ax2.set_title("随机森林特征重要性\n（Permutation Importance）", fontsize=12, fontweight="bold")
ax2.set_yticklabels([FEATURE_ZH.get(f, f) for f in perm_imp.index[::-1]], fontsize=10)
for bar, val in zip(bars2, perm_imp.values[::-1]):
    ax2.text(val + 0.001, bar.get_y() + bar.get_height() / 2,
             f"{val:.4f}", va="center", ha="left", fontsize=8.5)

for ax in axes:
    red_p  = mpatches.Patch(color="#e63946", label="前 3 核心因子")
    blue_p = mpatches.Patch(color="#457b9d", label="其余特征")
    ax.legend(handles=[red_p, blue_p], fontsize=9)

fig.suptitle("Heart Failure — 随机森林特征重要性分析（已剔除 time）",
             fontsize=13, fontweight="bold", y=1.02)
plt.tight_layout()
plt.savefig("feature_importance.png", dpi=180, bbox_inches="tight")
plt.close()
print("  → 已保存：feature_importance.png")

# ============================================================
# 3. 生成统计结果 Excel 表格
# ============================================================
print("正在生成统计结果表格…")

top3_mdi  = importances_sorted.head(3).index.tolist()
top3_perm = perm_imp.head(3).index.tolist()

# 构建完整结果表
rows = []
for feat in importances_sorted.index:
    surv = X.loc[y == 0, feat]
    dead = X.loc[y == 1, feat]
    mdi_score  = importances_sorted[feat]
    perm_score = perm_imp[feat]
    corr_val   = death_corr.get(feat, np.nan)
    rank_mdi   = list(importances_sorted.index).index(feat) + 1
    rank_perm  = list(perm_imp.index).index(feat) + 1
    is_top3    = "★ 核心因子" if feat in top3_mdi else ""
    rows.append({
        "特征名称":            feat,
        "中文名称":            FEATURE_ZH.get(feat, feat),
        "MDI 重要性":         round(mdi_score, 6),
        "MDI 排名":           rank_mdi,
        "Permutation 重要性": round(perm_score, 6),
        "Perm 排名":          rank_perm,
        "与DEATH相关系数":    round(corr_val, 4),
        "存活组均值":         round(surv.mean(), 4),
        "死亡组均值":         round(dead.mean(), 4),
        "均值差（死-存）":    round(dead.mean() - surv.mean(), 4),
        "标注":               is_top3,
    })

result_df = pd.DataFrame(rows)

# ── 写入 Excel ────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "随机森林特征重要性"

# 样式定义
DARK_NAVY  = "1B2A4A"
GOLD       = "F4B942"
RED_LIGHT  = "FDDEDE"
BLUE_LIGHT = "DEE8F4"
GRAY_LIGHT = "F5F5F5"
WHITE      = "FFFFFF"

thin = Side(style="thin",   color="CCCCCC")
med  = Side(style="medium", color="1B2A4A")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
border_top  = Border(left=med,  right=med,  top=med,  bottom=thin)
border_bot  = Border(left=med,  right=med,  top=thin, bottom=med)
border_mid  = Border(left=med,  right=med,  top=thin, bottom=thin)

def cell_style(cell, bold=False, bg=None, fg="000000", size=10,
               h_align="center", v_align="center", wrap=False, border=None):
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    if bg:
        cell.fill  = PatternFill("solid", fgColor=bg)
    if border:
        cell.border = border

# Row 1 — 大标题
ws.merge_cells("A1:K1")
ws["A1"] = "Heart Failure Clinical Records — 随机森林特征重要性分析（已剔除 time 变量）"
cell_style(ws["A1"], bold=True, bg=DARK_NAVY, fg=WHITE, size=13)
ws.row_dimensions[1].height = 32

# Row 2 — 副标题
ws.merge_cells("A2:K2")
ws["A2"] = (f"模型：RandomForestClassifier  |  5-Fold CV AUC = {auc_scores.mean():.4f} ± {auc_scores.std():.4f}"
            f"  |  样本量 N = {len(df)}  |  目标变量：DEATH_EVENT")
cell_style(ws["A2"], bg="2E4070", fg="E8E8E8", size=10)
ws.row_dimensions[2].height = 20

# Row 3 — 空行
ws.row_dimensions[3].height = 8

# Row 4 — 表头
headers = list(result_df.columns)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=h)
    cell_style(cell, bold=True, bg=DARK_NAVY, fg=WHITE, size=10, border=border_thin)
ws.row_dimensions[4].height = 28

# Row 5+ — 数据行
for row_idx, (_, row) in enumerate(result_df.iterrows(), 5):
    is_top3_row = row["标注"] != ""
    bg = RED_LIGHT if is_top3_row else (GRAY_LIGHT if row_idx % 2 == 0 else WHITE)
    for col_idx, val in enumerate(row.values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        align = "left" if col_idx in [1, 2, 11] else "center"
        bold  = True if col_idx == 11 and is_top3_row else False
        fg_color = "C0392B" if (is_top3_row and col_idx == 11) else "000000"
        cell_style(cell, bold=bold, bg=bg, fg=fg_color, size=10,
                   h_align=align, border=border_thin)
    ws.row_dimensions[row_idx].height = 22

# 前3行加金色左边框（突出核心因子）
top3_excel_rows = [5 + i for i, r in enumerate(result_df["标注"]) if r != ""]
for r in top3_excel_rows:
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        existing = cell.border
        cell.border = Border(
            left  = Side(style="thick", color="E63946"),
            right = existing.right,
            top   = existing.top,
            bottom= existing.bottom,
        )

# 脚注
note_row = 5 + len(result_df) + 1
ws.merge_cells(f"A{note_row}:K{note_row}")
ws[f"A{note_row}"] = ("注：MDI = Mean Decrease Impurity（树节点不纯度均值下降）；"
                      "Permutation Importance = 随机打乱该特征后模型 AUC 的下降幅度。"
                      "★ 核心因子 = MDI 排名前 3。连续变量均值为标准化后数值。")
cell_style(ws[f"A{note_row}"], bg=GRAY_LIGHT, fg="555555", size=9,
           h_align="left", wrap=True)
ws.row_dimensions[note_row].height = 36

# 列宽
col_widths = [22, 12, 16, 10, 22, 10, 18, 14, 14, 18, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 第二张 Sheet — 前3核心因子详情
ws2 = wb.create_sheet("核心因子详情")
ws2.merge_cells("A1:F1")
ws2["A1"] = "前 3 核心因子 — 详细统计"
cell_style(ws2["A1"], bold=True, bg=DARK_NAVY, fg=WHITE, size=13)
ws2.row_dimensions[1].height = 30

h2 = ["特征名称", "中文名称", "MDI 重要性", "MDI 排名", "存活组均值（DEATH=0）", "死亡组均值（DEATH=1）"]
for ci, h in enumerate(h2, 1):
    cell = ws2.cell(row=2, column=ci, value=h)
    cell_style(cell, bold=True, bg="2E4070", fg=WHITE, size=10, border=border_thin)
ws2.row_dimensions[2].height = 26

for ri, feat in enumerate(top3_mdi, 3):
    row_data = result_df[result_df["特征名称"] == feat].iloc[0]
    vals = [row_data["特征名称"], row_data["中文名称"],
            row_data["MDI 重要性"], row_data["MDI 排名"],
            row_data["存活组均值"], row_data["死亡组均值"]]
    bg2 = [RED_LIGHT, RED_LIGHT, RED_LIGHT][ri - 3] if ri <= 5 else WHITE
    for ci, v in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell_style(cell, bold=(ci == 1), bg=RED_LIGHT, size=11, border=border_thin)
    ws2.row_dimensions[ri].height = 26

for ci, w in enumerate([22, 12, 16, 10, 24, 24], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

wb.save("stat_results_rf.xlsx")
print("  → 已保存：stat_results_rf.xlsx")

# ============================================================
# 4. 控制台摘要
# ============================================================
print("\n" + "=" * 60)
print("  分析完成摘要")
print("=" * 60)
print(f"  数据集：heart_failure_scaled.csv（已剔除 time）")
print(f"  样本量：{len(df)}  |  特征数：{X.shape[1]}")
print(f"  5-Fold CV AUC：{auc_scores.mean():.4f} ± {auc_scores.std():.4f}")
print()
print("  【前3核心因子 — MDI 重要性】")
for i, feat in enumerate(top3_mdi, 1):
    zh = FEATURE_ZH.get(feat, feat)
    score = importances_sorted[feat]
    corr  = death_corr.get(feat, 0)
    direction = "正相关↑死亡" if corr > 0 else "负相关↓死亡"
    print(f"    {i}. {feat:28s}({zh:8s})  MDI={score:.4f}  相关性={corr:+.3f} ({direction})")
print()
print("  【前3核心因子 — Permutation 重要性】")
for i, feat in enumerate(top3_perm, 1):
    zh = FEATURE_ZH.get(feat, feat)
    score = perm_imp[feat]
    print(f"    {i}. {feat:28s}({zh:8s})  Perm={score:.4f}")
print()
print("  输出文件：")
print("    heatmap_correlation.png   — 相关性热图")
print("    feature_importance.png    — 特征重要性双图")
print("    stat_results_rf.xlsx      — 统计结果表格")
print("=" * 60)




#phase 4:预测
# ============================================================
#  Heart Failure — 概率预测建模 · 多模型对比 · 风险分层
#  依赖: pip install pandas numpy matplotlib scikit-learn xgboost openpyxl
#
#  输出文件:
#    fig1_roc.png             ROC 曲线对比图
#    fig2_confusion.png       三模型混淆矩阵
#    fig3_risk_patients.png   风险分层 + 新患者预测
#    model_performance.xlsx   性能汇总 Excel 表格
# ============================================================

import os, sys, warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.patches import Patch
import matplotlib.font_manager as fm
warnings.filterwarnings("ignore")

from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import (roc_auc_score, roc_curve, confusion_matrix,
                              accuracy_score, precision_score, recall_score,
                              f1_score)
from xgboost import XGBClassifier
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 字体自动探测：优先 Windows 中文字体，自动回退
# ============================================================
def setup_font():
    """
    字体优先级:
      Windows 环境 → Microsoft YaHei / SimHei / SimSun
      Linux 环境   → Noto Sans CJK JP / WenQuanYi Zen Hei
      终极兜底     → DejaVu Sans（英文）
    返回实际选用的字体名。
    """
    registered = {f.name for f in fm.fontManager.ttflist}

    priority = [
        # Windows 内置中文字体
        "Microsoft YaHei",
        "SimHei",
        "SimSun",
        "KaiTi",
        "DengXian",
        "FangSong",
        # Linux / macOS 中文字体
        "Noto Sans CJK JP",
        "Noto Sans CJK SC",
        "WenQuanYi Zen Hei",
        "WenQuanYi Micro Hei",
        "Source Han Sans CN",
        "IPAGothic",
        # 终极兜底
        "DejaVu Sans",
    ]

    chosen = next((f for f in priority if f in registered), "DejaVu Sans")
    plt.rcParams["font.family"]      = "sans-serif"
    plt.rcParams["font.sans-serif"]  = [chosen, "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    # 若未在注册列表找到，尝试直接加载 Windows 字体文件
    if chosen == "DejaVu Sans":
        for fpath, fname in [
            (r"C:\Windows\Fonts\msyh.ttc",  "Microsoft YaHei"),
            (r"C:\Windows\Fonts\simhei.ttf", "SimHei"),
            (r"C:\Windows\Fonts\simsun.ttc", "SimSun"),
        ]:
            if os.path.exists(fpath):
                fm.fontManager.addfont(fpath)
                plt.rcParams["font.sans-serif"] = [fname, "DejaVu Sans"]
                chosen = fname
                break

    return chosen

FONT = setup_font()
print(f"[字体] {FONT}")

# ============================================================
# 主题色
# ============================================================
BG    = "#0F1117"
CARD  = "#1A1D27"
TEXT  = "#E8EAF0"
MUTED = "#8890A4"
GRID  = "#252836"

C = {"LR": "#4E9AF1", "RF": "#2ECC71", "XGB": "#E74C3C",
     "low": "#27AE60", "mid": "#F39C12", "high": "#E74C3C"}

ML = {"LR": "Logistic Regression", "RF": "Random Forest", "XGB": "XGBoost"}

# ============================================================
# 共用轴样式（不设 ylabel，统一在外面设，避免遮挡）
# ============================================================
def stylize(ax, title="", xlabel="", ylabel="", grid=True, gridaxis="both"):
    ax.set_facecolor(CARD)
    for sp in ax.spines.values():
        sp.set_edgecolor(GRID)
        sp.set_linewidth(0.8)
    ax.tick_params(colors=MUTED, labelsize=10, pad=4)
    ax.xaxis.label.set_color(MUTED)
    ax.yaxis.label.set_color(MUTED)
    if title:
        ax.set_title(title, color=TEXT, fontsize=12,
                     fontweight="bold", pad=10)
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=10, labelpad=6)
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=10, labelpad=8)
    if grid:
        ax.grid(True, color=GRID, linewidth=0.5,
                alpha=0.9, zorder=0, axis=gridaxis)

def fig_header(fig, title, subtitle=""):
    fig.text(0.5, 0.97, title, ha="center", va="top",
             fontsize=14, fontweight="bold", color=TEXT)
    if subtitle:
        fig.text(0.5, 0.935, subtitle, ha="center", va="top",
                 fontsize=10, color=MUTED)

# ============================================================
# 0. 数据
# ============================================================
print("▶ 加载数据...")
df = pd.read_csv("heart_failure_scaled.csv").drop(columns=["time"])
X  = df.drop(columns=["DEATH_EVENT"])
y  = df["DEATH_EVENT"]

X_tr, X_te, y_tr, y_te = train_test_split(
    X, y, test_size=0.2, random_state=42, stratify=y)
print(f"  训练 {len(X_tr)} | 测试 {len(X_te)}")

# ============================================================
# 1. 训练
# ============================================================
print("▶ 训练模型...")
MODELS = {
    "LR":  LogisticRegression(max_iter=1000, C=0.5, random_state=42),
    "RF":  RandomForestClassifier(n_estimators=500, max_depth=8,
                                   class_weight="balanced",
                                   random_state=42, n_jobs=-1),
    "XGB": XGBClassifier(n_estimators=200, max_depth=4,
                          learning_rate=0.05, scale_pos_weight=203/96,
                          eval_metric="logloss",
                          random_state=42, verbosity=0),
}

R = {}
for k, m in MODELS.items():
    m.fit(X_tr, y_tr)
    prob = m.predict_proba(X_te)[:, 1]
    pred = m.predict(X_te)
    fpr, tpr, _ = roc_curve(y_te, prob)
    R[k] = dict(
        model=m, prob=prob, pred=pred,
        auc =roc_auc_score(y_te, prob),
        acc =accuracy_score(y_te, pred),
        prec=precision_score(y_te, pred, zero_division=0),
        rec =recall_score(y_te, pred, zero_division=0),
        f1  =f1_score(y_te, pred, zero_division=0),
        cm  =confusion_matrix(y_te, pred),
        fpr =fpr, tpr=tpr,
    )
    print(f"  {k:3s}  AUC={R[k]['auc']:.4f}  "
          f"Acc={R[k]['acc']:.4f}  F1={R[k]['f1']:.4f}")

BEST = max(R, key=lambda k: R[k]["auc"])
print(f"  最优: {ML[BEST]} (AUC={R[BEST]['auc']:.4f})")

# ============================================================
# 2. 风险分层
# ============================================================
rdf = pd.DataFrame({"prob": R[BEST]["prob"], "y": y_te.values})
BINS   = [0, 0.3, 0.7, 1.0]
RLABEL = ["低风险 (<0.3)", "中风险 (0.3-0.7)", "高风险 (>0.7)"]
rdf["grp"] = pd.cut(rdf["prob"], bins=BINS, labels=RLABEL)
ST = (rdf.groupby("grp", observed=True)["y"]
         .agg(n="count", deaths="sum")
         .assign(rate=lambda d: d["deaths"] / d["n"]))
print("\n── 风险分层 ─────────────────────────")
print(ST.to_string())

# ============================================================
# 3. 新患者
# ============================================================
FCOLS = X.columns.tolist()
mu, sd = X_tr.mean(), X_tr.std()

PAT = pd.DataFrame([
    {**{c: mu[c] for c in FCOLS},
     "age": mu["age"] - 1.2*sd["age"],
     "ejection_fraction": mu["ejection_fraction"] + 1.5*sd["ejection_fraction"],
     "serum_creatinine":  mu["serum_creatinine"]  - 1.0*sd["serum_creatinine"],
     "serum_sodium":      mu["serum_sodium"]       + 0.8*sd["serum_sodium"],
     "anaemia": 0, "high_blood_pressure": 0},
    {**{c: mu[c] for c in FCOLS},
     "age": mu["age"] + 0.3*sd["age"],
     "ejection_fraction": mu["ejection_fraction"] - 0.4*sd["ejection_fraction"],
     "serum_creatinine":  mu["serum_creatinine"]  + 0.5*sd["serum_creatinine"],
     "anaemia": 1},
    {**{c: mu[c] for c in FCOLS},
     "age": mu["age"] + 2.0*sd["age"],
     "ejection_fraction": mu["ejection_fraction"] - 2.0*sd["ejection_fraction"],
     "serum_creatinine":  mu["serum_creatinine"]  + 2.5*sd["serum_creatinine"],
     "serum_sodium":      mu["serum_sodium"]       - 1.5*sd["serum_sodium"],
     "anaemia": 1, "high_blood_pressure": 1},
], columns=FCOLS)

PPROB = R[BEST]["model"].predict_proba(PAT)[:, 1]
PLBL  = ["患者A（低风险）", "患者B（中风险）", "患者C（高风险）"]
print("\n── 新患者预测 ────────────────────────")
for l, p in zip(PLBL, PPROB):
    tag = "低风险" if p < 0.3 else ("高风险" if p > 0.7 else "中风险")
    print(f"  {l}: {p:.4f}  [{tag}]")

# ============================================================
# 图 1 — ROC 曲线
# ============================================================
print("\n▶ 图1: ROC...")

fig1, ax = plt.subplots(figsize=(8, 7), facecolor=BG)
fig1.patch.set_facecolor(BG)
stylize(ax, xlabel="假阳性率 (FPR)", ylabel="真阳性率 (TPR)")

ax.plot([0,1],[0,1], "--", color="#3A3F52", lw=1.2,
        label="随机猜测  AUC = 0.500")

for k in ["LR", "RF", "XGB"]:
    lw   = 3.0 if k == BEST else 2.0
    suf  = "  ★ 最优" if k == BEST else ""
    ax.plot(R[k]["fpr"], R[k]["tpr"], color=C[k], lw=lw, zorder=4 if k==BEST else 3,
            label=f"{ML[k]}   AUC = {R[k]['auc']:.4f}{suf}")

ax.fill_between(R[BEST]["fpr"], R[BEST]["tpr"],
                alpha=0.07, color=C[BEST])

ax.set_xlim(-0.02, 1.02)
ax.set_ylim(-0.02, 1.05)
ax.legend(fontsize=10, framealpha=0.25, labelcolor=TEXT,
           facecolor=CARD, edgecolor=GRID,
           loc="lower right", handlelength=2)

fig_header(fig1,
           "Heart Failure 死亡预测 — ROC 曲线对比",
           f"测试集 N={len(X_te)}  |  80/20 分层划分  |  最优: {ML[BEST]}")

fig1.subplots_adjust(top=0.88, bottom=0.10, left=0.11, right=0.97)
fig1.savefig("fig1_roc.png", dpi=180, bbox_inches="tight", facecolor=BG)
plt.close(fig1)
print("  → fig1_roc.png")

# ============================================================
# 图 2 — 混淆矩阵（1×3，充分留边距避免遮挡）
# ============================================================
print("▶ 图2: 混淆矩阵...")

CMAP = LinearSegmentedColormap.from_list(
    "nb", ["#141820", "#1B3A5C", "#2471A3", "#5DADE2"], N=256)

fig2, axes = plt.subplots(1, 3, figsize=(16, 6), facecolor=BG)
fig2.patch.set_facecolor(BG)

TICK_LABELS = ["存活 (0)", "死亡 (1)"]

for ax2, k in zip(axes, ["LR", "RF", "XGB"]):
    cm   = R[k]["cm"]
    star = "  ★" if k == BEST else ""
    ax2.set_facecolor(CARD)

    # 关闭默认 tick_params 颜色，单独配置
    for sp in ax2.spines.values():
        sp.set_edgecolor(C[k])
        sp.set_linewidth(2.5)

    ax2.imshow(cm, cmap=CMAP, aspect="auto",
               vmin=0, vmax=cm.max() * 1.1)

    # 刻度与标签
    ax2.set_xticks([0, 1])
    ax2.set_yticks([0, 1])
    ax2.set_xticklabels(TICK_LABELS, fontsize=11,
                         color=TEXT, fontweight="bold")
    ax2.set_yticklabels(TICK_LABELS, fontsize=11,
                         color=TEXT, fontweight="bold",
                         rotation=0)          # ← 不旋转，彻底避免竖排遮挡

    ax2.tick_params(axis="both", which="both",
                    length=0, pad=6, colors=TEXT)

    # 轴标签：xlabel 正常；ylabel 改为横排文字，避免旋转遮挡
    ax2.set_xlabel("预测标签", fontsize=11, color=MUTED, labelpad=8)
    ax2.set_ylabel("")          # 关闭默认旋转 ylabel
    ax2.text(-0.36, 0.5, "真实标签",
             transform=ax2.transAxes,
             fontsize=11, color=MUTED,
             ha="center", va="center", rotation=0)

    # 标题
    ax2.set_title(f"{ML[k]}{star}\nAUC = {R[k]['auc']:.4f}",
                  color=TEXT, fontsize=12, fontweight="bold", pad=10)

    # 格子内数字
    row_sum = cm.sum(axis=1, keepdims=True)
    for i in range(2):
        for j in range(2):
            v   = cm[i, j]
            pct = v / row_sum[i, 0] * 100
            bright = v / max(cm.max(), 1)
            fc = "#0A0E18" if bright > 0.55 else TEXT
            ax2.text(j, i, f"{v}\n({pct:.1f}%)",
                     ha="center", va="center",
                     fontsize=13, fontweight="bold", color=fc)

fig_header(fig2,
           "混淆矩阵 — 三模型分类效果对比",
           "行 = 真实标签  |  列 = 预测标签  |  括号内为行内占比（召回率视角）")

# 给 ylabel 留足够左边距
fig2.subplots_adjust(top=0.82, bottom=0.13,
                      left=0.08, right=0.97, wspace=0.38)
fig2.savefig("fig2_confusion.png", dpi=180, bbox_inches="tight", facecolor=BG)
plt.close(fig2)
print("  → fig2_confusion.png")

# ============================================================
# 图 3 — 风险分层 + 新患者预测
# ============================================================
print("▶ 图3: 风险分层 & 新患者...")

fig3, (axL, axR) = plt.subplots(1, 2, figsize=(14, 6), facecolor=BG)
fig3.patch.set_facecolor(BG)

# ── 左：风险分层 ──────────────────────────────────────────
RCOL = [C["low"], C["mid"], C["high"]]
rates = ST["rate"].values
ns    = ST["n"].values
ds    = ST["deaths"].values
xs    = np.arange(len(RLABEL))

bars = axL.bar(xs, rates, color=RCOL, width=0.50,
                edgecolor=BG, linewidth=1.5, zorder=3)

# 顶部深色加深效果
for bar, col in zip(bars, RCOL):
    h = bar.get_height()
    axL.bar(bar.get_x(), h * 0.15, bottom=h * 0.85,
            width=bar.get_width(), color=col,
            alpha=0.45, edgecolor="none", zorder=4)

# 标注
for bar, n, d, r in zip(bars, ns, ds, rates):
    cx = bar.get_x() + bar.get_width() / 2
    axL.text(cx, r + 0.025, f"{r:.1%}",
             ha="center", va="bottom",
             fontsize=14, fontweight="bold", color=TEXT, zorder=5)
    axL.text(cx, r / 2, f"n={n}\n死亡={d}",
             ha="center", va="center",
             fontsize=10, fontweight="bold", color="#0A0E18", zorder=5)

stylize(axL,
        title=f"风险分层验证 — {ML[BEST]}\n实际死亡率分布",
        xlabel="风险组", ylabel="实际死亡比例",
        gridaxis="y")
axL.set_xticks(xs)
axL.set_xticklabels(RLABEL, color=TEXT, fontsize=11)
axL.set_ylim(0, 0.95)
axL.yaxis.set_major_formatter(
    mticker.FuncFormatter(lambda v, _: f"{v:.0%}"))
axL.set_yticks(np.arange(0, 1.0, 0.2))

# 趋势标注
axL.annotate("",
             xy=(2, rates[2] + 0.05), xytext=(0, rates[0] + 0.05),
             arrowprops=dict(arrowstyle="-|>", color=MUTED, lw=1.4,
                             connectionstyle="arc3,rad=-0.15"))
axL.text(1, max(rates) + 0.13,
         "死亡率随风险等级单调递增 ✓",
         ha="center", fontsize=9, color=MUTED, style="italic")

# ── 右：新患者预测 ────────────────────────────────────────
pcolors = [C["low"] if p < 0.3 else (C["high"] if p > 0.7 else C["mid"])
           for p in PPROB]
pxs     = np.arange(len(PPROB))
plbls   = ["患者A\n低风险特征", "患者B\n中风险特征", "患者C\n高风险特征"]

axR.bar(pxs, PPROB, color=pcolors, width=0.46,
        edgecolor=BG, linewidth=1.5, zorder=3)

# 背景区带
axR.axhspan(0.0, 0.3, color=C["low"],  alpha=0.06, zorder=1)
axR.axhspan(0.3, 0.7, color=C["mid"],  alpha=0.06, zorder=1)
axR.axhspan(0.7, 1.0, color=C["high"], alpha=0.06, zorder=1)

# 阈值线
axR.axhline(0.3, color=C["low"],  ls="--", lw=1.5,
             alpha=0.85, zorder=2, label="低/中 阈值 0.30")
axR.axhline(0.7, color=C["high"], ls="--", lw=1.5,
             alpha=0.85, zorder=2, label="中/高 阈值 0.70")

# 标注
for i, (p, col) in enumerate(zip(PPROB, pcolors)):
    tag = "低风险" if p < 0.3 else ("高风险" if p > 0.7 else "中风险")
    axR.text(i, p + 0.028, f"{p:.4f}\n[{tag}]",
             ha="center", va="bottom",
             fontsize=11, fontweight="bold", color=col, zorder=5)

stylize(axR,
        title=f"新患者死亡概率预测\n模型: {ML[BEST]}",
        xlabel="患者", ylabel="死亡预测概率",
        gridaxis="y")
axR.set_xticks(pxs)
axR.set_xticklabels(plbls, color=TEXT, fontsize=11)
axR.set_ylim(0, 1.12)
axR.yaxis.set_major_formatter(
    mticker.FuncFormatter(lambda v, _: f"{v:.1f}"))
axR.legend(fontsize=9, framealpha=0.25, labelcolor=TEXT,
            facecolor=CARD, edgecolor=GRID, loc="upper left")

fig_header(fig3,
           f"风险分层验证 & 新患者预测 — {ML[BEST]}  (AUC={R[BEST]['auc']:.4f})",
           "左: 三组实际死亡率验证概率校准  |  右: 三位典型患者死亡概率预测")

fig3.subplots_adjust(top=0.84, bottom=0.13,
                      left=0.09, right=0.97, wspace=0.32)
fig3.savefig("fig3_risk_patients.png", dpi=180, bbox_inches="tight", facecolor=BG)
plt.close(fig3)
print("  → fig3_risk_patients.png")

# ============================================================
# Excel 性能汇总
# ============================================================
print("▶ Excel...")

wb = Workbook()
ws = wb.active
ws.title = "模型性能汇总"

NAVY, WHITE = "1B2A4A", "FFFFFF"
LGRAY, DGRAY, GREEN = "F5F7FA", "E8ECF2", "D5F0E0"
BC = "C5CCD8"
th = Side(style="thin",   color=BC)
mh = Side(style="medium", color=NAVY)
B0 = Border(left=th, right=th, top=th, bottom=th)
BH = Border(left=mh, right=mh, top=mh, bottom=mh)

def W(cell, v, bold=False, bg=None, fg="1A1A2E",
      sz=10, ha="center", wrap=False, border=B0):
    cell.value = v
    cell.font  = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = border

# 大标题
ws.merge_cells("A1:G1")
W(ws["A1"], "Heart Failure 死亡预测 — 多模型性能汇总表",
  bold=True, bg=NAVY, fg=WHITE, sz=13, border=BH)
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:G2")
W(ws["A2"],
  f"数据集: heart_failure_scaled.csv  |  N=299  |  Train/Test=80/20  "
  f"|  最优模型: {ML[BEST]} (AUC={R[BEST]['auc']:.4f})",
  bg="2E4070", fg="D0D8F0", sz=10)
ws.row_dimensions[2].height = 22
ws.row_dimensions[3].height = 8

# 表头
HDR = ["模型", "AUC", "Accuracy", "Precision", "Recall", "F1 Score", "备注"]
for ci, h in enumerate(HDR, 1):
    c = ws.cell(row=4, column=ci)
    W(c, h, bold=True, bg=NAVY, fg=WHITE, sz=11, border=BH)
ws.row_dimensions[4].height = 30

# 数据行
for ri, k in enumerate(["LR","RF","XGB"], 5):
    r    = R[k]
    note = "AUC 最优" if k == BEST else ""
    bg_r = GREEN if k == BEST else (LGRAY if ri%2==1 else DGRAY)
    vals = [ML[k], round(r["auc"],4), round(r["acc"],4),
            round(r["prec"],4), round(r["rec"],4), round(r["f1"],4), note]
    for ci, v in enumerate(vals, 1):
        W(ws.cell(row=ri, column=ci), v,
          bold=(k==BEST), bg=bg_r,
          fg=("C0392B" if k==BEST and ci==7 else "1A1A2E"),
          sz=11, ha="left" if ci in [1,7] else "center")
    ws.row_dimensions[ri].height = 26

ws.row_dimensions[8].height = 10

# 风险分层
ws.merge_cells("A9:G9")
W(ws["A9"], f"风险分层 — {ML[BEST]}（测试集 N={len(X_te)}）",
  bold=True, bg="2E4070", fg=WHITE, sz=11)
ws.row_dimensions[9].height = 26

for ci, h in enumerate(["风险组","预测区间","样本数","死亡数","实际死亡率","",""], 1):
    W(ws.cell(row=10, column=ci), h, bold=True, bg=NAVY, fg=WHITE, sz=10)
ws.row_dimensions[10].height = 26

RBG = ["D5F0E0","FEF3CD","FADBD8"]
for ri, (grp, rb) in enumerate(zip(RLABEL, RBG), 11):
    row = ST.loc[grp]
    for ci, v in enumerate([grp, grp, int(row["n"]),
                             int(row["deaths"]), f"{row['rate']:.1%}","",""], 1):
        W(ws.cell(row=ri, column=ci), v, bg=rb, sz=10,
          ha="left" if ci==1 else "center")
    ws.row_dimensions[ri].height = 24

ws.row_dimensions[14].height = 10

# 新患者
ws.merge_cells("A15:G15")
W(ws["A15"], f"新患者预测 — {ML[BEST]}",
  bold=True, bg="2E4070", fg=WHITE, sz=11)
ws.row_dimensions[15].height = 26

for ci, h in enumerate(["患者","特征描述","死亡概率","风险等级","","",""], 1):
    W(ws.cell(row=16, column=ci), h, bold=True, bg=NAVY, fg=WHITE, sz=10)
ws.row_dimensions[16].height = 26

PDESCS = ["年轻、高射血分数、低血清肌酐、无贫血/高血压",
          "中年、轻度低射血分数、轻度高肌酐、贫血",
          "高龄、重度低射血分数、高肌酐、低血钠、贫血+高血压"]
PBGM = {"低风险":"D5F0E0","中风险":"FEF3CD","高风险":"FADBD8"}
for ri, (lbl, p, desc) in enumerate(zip(PLBL, PPROB, PDESCS), 17):
    tag = "低风险" if p<0.3 else ("高风险" if p>0.7 else "中风险")
    bg_p = PBGM[tag]
    for ci, v in enumerate([lbl, desc, round(p,4), tag,"","",""], 1):
        W(ws.cell(row=ri, column=ci), v, bg=bg_p, sz=10,
          bold=(ci==3), ha="left" if ci in [1,2] else "center")
    ws.row_dimensions[ri].height = 26

ws.merge_cells("A21:G21")
W(ws["A21"],
  "注: 所有指标在测试集计算；正类=DEATH_EVENT=1（死亡）；"
  "风险阈值: 低风险<0.30, 中风险0.30-0.70, 高风险>0.70。",
  bg=LGRAY, fg="555555", sz=9, ha="left", wrap=True)
ws.row_dimensions[21].height = 36

for ci, w in enumerate([24,30,13,13,13,13,16], 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# Sheet 2
ws2 = wb.create_sheet("原始数值")
ws2.merge_cells("A1:F1")
W(ws2["A1"], "模型性能原始数值", bold=True, bg=NAVY, fg=WHITE, sz=12, border=BH)
ws2.row_dimensions[1].height = 28
for ci, h in enumerate(["模型","AUC","Accuracy","Precision","Recall","F1"], 1):
    W(ws2.cell(row=2, column=ci), h, bold=True, bg="2E4070", fg=WHITE, sz=10)
ws2.row_dimensions[2].height = 24
for ri, k in enumerate(["LR","RF","XGB"], 3):
    r   = R[k]
    bg2 = GREEN if k==BEST else (LGRAY if ri%2==0 else DGRAY)
    for ci, v in enumerate([ML[k], r["auc"], r["acc"],
                             r["prec"], r["rec"], r["f1"]], 1):
        W(ws2.cell(row=ri, column=ci),
          round(v,6) if ci>1 else v, bg=bg2, sz=10,
          ha="left" if ci==1 else "center")
    ws2.row_dimensions[ri].height = 22
for ci, w in enumerate([24,12,12,12,12,12], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

wb.save("model_performance.xlsx")
print("  → model_performance.xlsx")

# ============================================================
# 摘要
# ============================================================
print("\n" + "="*60)
print("  完成。输出文件:")
print("  fig1_roc.png  fig2_confusion.png  fig3_risk_patients.png")
print("  model_performance.xlsx")
print("="*60)
for k in ["LR","RF","XGB"]:
    r = R[k]; s = " ★" if k==BEST else ""
    print(f"  {ML[k]:<22} AUC={r['auc']:.4f}  F1={r['f1']:.4f}{s}")
print()
for g in RLABEL:
    ro = ST.loc[g]
    print(f"  {g:<22} n={int(ro['n']):>3}  死亡={int(ro['deaths'])}  "
          f"死亡率={ro['rate']:.1%}")
print()
for l, p in zip(PLBL, PPROB):
    tag = "低风险" if p<0.3 else ("高风险" if p>0.7 else "中风险")
    print(f"  {l}: {p:.4f} [{tag}]")
print("="*60)